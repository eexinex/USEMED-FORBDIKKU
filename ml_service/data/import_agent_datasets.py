#!/usr/bin/env python3
"""Convert agent-provided diabetes/hypertension Excel datasets into USE MED SQL.

The source workbooks are wide longitudinal data dictionaries:
one row per patient, with columns such as vitalsign_sbp_0 and lab_hba1c_12.
This importer normalizes them into patients, visits, patient_longitudinal_records,
and ML feature/label snapshots that the XGBoost training pipeline can consume.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DATASET_DIR = ROOT / ".agents" / "dataset_for_agent"
DEFAULT_OUTPUT = ROOT / "backend" / "database" / "agent_dataset_seed.sql"
DEFAULT_CHUNK_DIR = ROOT / "backend" / "database" / "agent_dataset_seed_chunks"
BASELINE_DATE = date(2026, 1, 1)


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    return value


def as_float(value: Any) -> float | None:
    value = clean_value(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def as_int(value: Any) -> int | None:
    number = as_float(value)
    if number is None:
        return None
    return int(round(number))


def bounded_float(value: Any, minimum: float | None = None, maximum: float | None = None) -> float | None:
    number = as_float(value)
    if number is None:
        return None
    if minimum is not None and number < minimum:
        return None
    if maximum is not None and number > maximum:
        return None
    return number


def hba1c_value(value: Any) -> float | None:
    return bounded_float(value, 0, 30)


def bmi_value(value: Any) -> float | None:
    return bounded_float(value, 0, 100)


def weight_value(value: Any) -> float | None:
    return bounded_float(value, 0, 500)


def height_value(value: Any) -> float | None:
    return bounded_float(value, 0, 250)


def sbp_value(value: Any) -> int | None:
    number = as_int(value)
    return number if number is not None and 40 <= number <= 300 else None


def dbp_value(value: Any) -> int | None:
    number = as_int(value)
    return number if number is not None and 20 <= number <= 200 else None


def pulse_value(value: Any) -> int | None:
    number = as_int(value)
    return number if number is not None and 20 <= number <= 250 else None


def respiratory_value(value: Any) -> int | None:
    number = as_int(value)
    return number if number is not None and 4 <= number <= 80 else None


def oxygen_value(value: Any) -> int | None:
    number = as_int(value)
    return number if number is not None and 0 <= number <= 100 else None


def flag(value: Any) -> int:
    number = as_float(value)
    return 1 if number is not None and number >= 0.5 else 0


def sql_literal(value: Any) -> str:
    value = clean_value(value)
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, date):
        return "'" + value.isoformat() + "'"
    text = str(value).replace("'", "''")
    return "'" + text + "'"


def insert_sql(table: str, data: dict[str, Any], conflict: str = "") -> str:
    columns = list(data.keys())
    values = ", ".join(sql_literal(data[col]) for col in columns)
    sql = f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({values})"
    if conflict:
        sql += " " + conflict
    return sql + ";"


def iter_workbook_rows(path: Path) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["data"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows)]
    for row in rows:
        yield {headers[i]: clean_value(row[i] if i < len(row) else None) for i in range(len(headers))}


def period_columns(row: dict[str, Any]) -> dict[int, dict[str, Any]]:
    periods: dict[int, dict[str, Any]] = {}
    for key, value in row.items():
        if value is None:
            continue
        match = re.match(r"(.+)_(-?\d+)$", key)
        if not match:
            continue
        feature = match.group(1)
        period = int(match.group(2))
        periods.setdefault(period, {})[feature] = value
    return periods


def current_or_nearest(periods: dict[int, dict[str, Any]], feature: str, baseline: int = 0) -> Any:
    choices = []
    for period, values in periods.items():
        if feature in values and values[feature] is not None:
            choices.append((abs(period - baseline), period, values[feature]))
    if not choices:
        return None
    return sorted(choices, key=lambda item: (item[0], abs(item[1])))[0][2]


def future_value(periods: dict[int, dict[str, Any]], feature: str, min_period: int = 1, max_period: int = 69) -> Any:
    choices = []
    for period, values in periods.items():
        if min_period <= period <= max_period and feature in values and values[feature] is not None:
            choices.append((period, values[feature]))
    if not choices:
        return None
    return sorted(choices, key=lambda item: item[0])[0][1]


def trend(periods: dict[int, dict[str, Any]], feature: str) -> float | None:
    current = as_float(current_or_nearest(periods, feature, 0))
    previous = None
    for period in sorted((p for p in periods if p < 0), reverse=True):
        value = as_float(periods[period].get(feature))
        if value is not None:
            previous = value
            break
    if current is None or previous is None:
        return None
    return round(current - previous, 4)


def disease_for(row: dict[str, Any], disease_group: str) -> str:
    diseases: list[str] = []
    if disease_group == "diabetes":
        if flag(row.get("type1")):
            diseases.append("Type 1 Diabetes Mellitus")
        if flag(row.get("type2")):
            diseases.append("Type 2 Diabetes Mellitus")
        if flag(row.get("gdm")):
            diseases.append("Gestational Diabetes Mellitus")
        if flag(row.get("co_ht")):
            diseases.append("Hypertension")
    else:
        diseases.append("Hypertension")
        if flag(row.get("co_dm")):
            diseases.append("Diabetes Mellitus")
    for key, label in [
        ("co_ckd", "CKD"),
        ("co_stroke", "Stroke"),
        ("co_hf", "Heart Failure"),
        ("co_cad", "CAD"),
        ("co_arrhythmias", "Arrhythmia"),
        ("co_atrial_fibrillation", "Atrial Fibrillation"),
        ("co_dementia", "Dementia"),
    ]:
        if flag(row.get(key)):
            diseases.append(label)
    return " / ".join(dict.fromkeys(diseases)) or disease_group.title()


def medication_summary(row: dict[str, Any]) -> str:
    meds = [key.replace("med_", "") for key, value in row.items() if key.startswith("med_") and not re.search(r"_-?\d+$", key) and flag(value)]
    return ", ".join(meds)


def risk_from_labels(disease_group: str, labels: dict[str, Any], row: dict[str, Any]) -> tuple[int, str, int]:
    score = 20
    high_watch = 0
    age = as_int(row.get("age")) or 0
    if age >= 75:
        score += 15
    elif age >= 60:
        score += 8
    for key in ["co_ckd", "co_stroke", "co_hf", "co_cad"]:
        if flag(row.get(key)):
            score += 8
    if disease_group == "diabetes":
        future_hba1c = as_float(labels.get("future_hba1c"))
        future_fpg = as_float(labels.get("future_fpg"))
        if future_hba1c is not None:
            if future_hba1c >= 9:
                score += 35
            elif future_hba1c >= 8:
                score += 24
            elif future_hba1c >= 7:
                score += 12
        if future_fpg is not None and future_fpg >= 180:
            score += 12
    else:
        future_sbp = as_float(labels.get("future_sbp"))
        future_dbp = as_float(labels.get("future_dbp"))
        if future_sbp is not None:
            if future_sbp >= 160:
                score += 35
            elif future_sbp >= 140:
                score += 22
        if future_dbp is not None and future_dbp >= 90:
            score += 10
    score = max(0, min(100, score))
    if score >= 80:
        level = "High"
        high_watch = 1
    elif score >= 55:
        level = "Medium"
    else:
        level = "Low"
    return score, level, high_watch


def build_features(row: dict[str, Any], periods: dict[int, dict[str, Any]], disease_group: str) -> dict[str, Any]:
    base = {
        "age": as_int(row.get("age")),
        "sex": row.get("sex"),
        "identify_by": row.get("identify_by"),
        "disease_group": disease_group,
        "current_sbp": sbp_value(current_or_nearest(periods, "vitalsign_sbp")),
        "current_dbp": dbp_value(current_or_nearest(periods, "vitalsign_dbp")),
        "current_hr": pulse_value(current_or_nearest(periods, "vitalsign_hr")),
        "current_bmi": bmi_value(current_or_nearest(periods, "vitalsign_bmi")),
        "current_fpg": as_float(current_or_nearest(periods, "lab_fpg")),
        "current_hba1c": hba1c_value(current_or_nearest(periods, "lab_hba1c")),
        "current_ldl": as_float(current_or_nearest(periods, "lab_ldl")),
        "current_chol": as_float(current_or_nearest(periods, "lab_chol")),
        "trend_sbp": trend(periods, "vitalsign_sbp"),
        "trend_dbp": trend(periods, "vitalsign_dbp"),
        "trend_hba1c": trend(periods, "lab_hba1c"),
        "trend_fpg": trend(periods, "lab_fpg"),
    }
    for key, value in row.items():
        if re.search(r"_-?\d+$", key):
            continue
        if key.startswith("co_") or key.startswith("med_") or key in ["type1", "type2", "gdm"]:
            base[key] = flag(value)
    return base


def build_labels(periods: dict[int, dict[str, Any]], disease_group: str) -> dict[str, Any]:
    labels = {
        "future_sbp": as_float(future_value(periods, "vitalsign_sbp")),
        "future_dbp": as_float(future_value(periods, "vitalsign_dbp")),
        "future_hba1c": hba1c_value(future_value(periods, "lab_hba1c")),
        "future_fpg": as_float(future_value(periods, "lab_fpg")),
    }
    labels["hba1c_uncontrolled"] = 1 if (labels["future_hba1c"] is not None and labels["future_hba1c"] >= 8.0) else 0
    labels["bp_uncontrolled"] = 1 if (
        (labels["future_sbp"] is not None and labels["future_sbp"] >= 140.0)
        or (labels["future_dbp"] is not None and labels["future_dbp"] >= 90.0)
    ) else 0
    labels["needs_followup"] = 1 if labels["hba1c_uncontrolled"] or labels["bp_uncontrolled"] else 0
    return labels


def period_date(period: int) -> date:
    return BASELINE_DATE + timedelta(days=period)


def visit_title(disease_group: str, period: int) -> str:
    if period == 0:
        when = "baseline"
    elif period > 0:
        when = f"+{period} days"
    else:
        when = f"{period} days"
    return f"{disease_group.title()} longitudinal visit ({when})"


def build_sql(dataset_dir: Path) -> str:
    ml_schema = (ROOT / "backend" / "database" / "ml_schema.sql").read_text(encoding="utf-8")
    statements: list[str] = [
        "-- Generated from .agents/dataset_for_agent by ml_service/data/import_agent_datasets.py",
        ml_schema.strip(),
        "START TRANSACTION;",
    ]
    files = [
        ("diabetes", dataset_dir / "data_dictionary_diabetes_example.xlsx", "DM"),
        ("hypertension", dataset_dir / "data_dictionary_hypertension_example.xlsx", "HT"),
    ]
    for disease_group, path, prefix in files:
        source_dataset = path.name
        for idx, row in enumerate(iter_workbook_rows(path), 1):
            periods = period_columns(row)
            features = build_features(row, periods, disease_group)
            labels = build_labels(periods, disease_group)
            score, level, high_watch = risk_from_labels(disease_group, labels, row)
            hn = f"{prefix}{idx:04d}"
            sex = str(row.get("sex") or "").upper()
            gender = "หญิง" if sex == "FEMALE" else "ชาย" if sex == "MALE" else None
            patient_name = f"{disease_group.title()} Patient {idx}"
            disease = disease_for(row, disease_group)
            statements.append(insert_sql(
                "patients",
                {
                    "hn": hn,
                    "password": "123456",
                    "full_name": patient_name,
                    "gender": gender,
                    "age": as_int(row.get("age")),
                    "disease": disease,
                    "care_area": "OPD",
                    "hospital": "USE MED Dataset Hospital",
                    "ward": f"{disease_group.title()} Clinic",
                    "high_watch": high_watch,
                    "department": "อายุรกรรม",
                    "risk_level": level,
                    "risk_score": score,
                    "additional_medication": medication_summary(row),
                    "registration_source": "agent_dataset",
                    "registration_status": "active",
                },
                "ON CONFLICT (hn) DO UPDATE SET full_name=EXCLUDED.full_name, gender=EXCLUDED.gender, age=EXCLUDED.age, disease=EXCLUDED.disease, care_area=EXCLUDED.care_area, hospital=EXCLUDED.hospital, ward=EXCLUDED.ward, high_watch=EXCLUDED.high_watch, department=EXCLUDED.department, risk_level=EXCLUDED.risk_level, risk_score=EXCLUDED.risk_score, additional_medication=EXCLUDED.additional_medication, registration_source=EXCLUDED.registration_source, registration_status=EXCLUDED.registration_status",
            ))
            statements.append(
                "INSERT INTO ml_patient_features "
                "(patient_id, hn, disease_group, source_dataset, baseline_period, feature_snapshot, label_snapshot) "
                f"VALUES ((SELECT id FROM patients WHERE hn = {sql_literal(hn)} LIMIT 1), "
                f"{sql_literal(hn)}, {sql_literal(disease_group)}, {sql_literal(source_dataset)}, 0, "
                f"{sql_literal(json.dumps(features, ensure_ascii=False, separators=(',', ':')))}, "
                f"{sql_literal(json.dumps(labels, ensure_ascii=False, separators=(',', ':')))}"
                ") ON CONFLICT (hn, disease_group) DO UPDATE SET "
                "patient_id=EXCLUDED.patient_id, source_dataset=EXCLUDED.source_dataset, "
                "feature_snapshot=EXCLUDED.feature_snapshot, label_snapshot=EXCLUDED.label_snapshot;"
            )
            statements.append(insert_sql(
                "ml_training_snapshots",
                {
                    "source_dataset": source_dataset,
                    "hn": hn,
                    "disease_group": disease_group,
                    "feature_snapshot": json.dumps(features, ensure_ascii=False, separators=(",", ":")),
                    "label_snapshot": json.dumps(labels, ensure_ascii=False, separators=(",", ":")),
                    "split_name": None,
                },
            ))
            useful_periods = sorted(
                p for p, values in periods.items()
                if any(k in values for k in ["vitalsign_sbp", "vitalsign_dbp", "vitalsign_hr", "vitalsign_bmi", "lab_hba1c", "lab_fpg", "lab_chol", "lab_ldl"])
            )
            # Keep the seed practical for the app: dense clinical timeline near baseline plus future labels.
            selected_periods = [p for p in useful_periods if -7 <= p <= 30]
            if 0 not in selected_periods and 0 in useful_periods:
                selected_periods.append(0)
            if not selected_periods:
                selected_periods = useful_periods[:5]
            for period in sorted(set(selected_periods)):
                values = periods[period]
                statements.append(
                    "INSERT INTO visits (patient_id, doctor_id, visit_date, title, diagnosis, treatment_plan, "
                    "systolic, diastolic, pulse, glucose, hba1c, bmi, cholesterol, visit_type, care_area, hospital, "
                    "weight_kg, height_cm, respiratory_rate, oxygen_saturation, current_medications, risk_score, risk_level) "
                    f"SELECT p.id, d.id, {sql_literal(period_date(period))}, {sql_literal(visit_title(disease_group, period))}, "
                    f"{sql_literal(disease)}, {sql_literal('Imported longitudinal dataset visit')}, "
                    f"{sql_literal(sbp_value(values.get('vitalsign_sbp')))}, {sql_literal(dbp_value(values.get('vitalsign_dbp')))}, "
                    f"{sql_literal(pulse_value(values.get('vitalsign_hr')))}, {sql_literal(as_float(values.get('lab_fpg')))}, "
                    f"{sql_literal(hba1c_value(values.get('lab_hba1c')))}, {sql_literal(bmi_value(values.get('vitalsign_bmi')))}, "
                    f"{sql_literal(as_float(values.get('lab_chol')))}, {sql_literal('Dataset')}, {sql_literal('OPD')}, "
                    f"{sql_literal('USE MED Dataset Hospital')}, {sql_literal(weight_value(values.get('vitalsign_wt')))}, "
                    f"{sql_literal(height_value(values.get('vitalsign_ht')))}, {sql_literal(respiratory_value(values.get('vitalsign_resp')))}, "
                    f"{sql_literal(oxygen_value(values.get('vitalsign_o2sat')))}, {sql_literal(medication_summary(row))}, "
                    f"{sql_literal(score)}, {sql_literal(level)} "
                    f"FROM patients p LEFT JOIN doctors d ON d.username='doctor1' "
                    f"WHERE p.hn={sql_literal(hn)} AND NOT EXISTS ("
                    f"SELECT 1 FROM visits v WHERE v.patient_id=p.id AND v.visit_date={sql_literal(period_date(period))} AND v.title={sql_literal(visit_title(disease_group, period))});"
                )
                statements.append(
                    "INSERT INTO patient_longitudinal_records "
                    "(patient_id, hn, period_offset, systolic, diastolic, pulse, respiratory_rate, hba1c, c_peptide, lipid_ldl, "
                    "med_metformin, med_insulin, med_arb, med_ccb, med_acei) "
                    f"SELECT p.id, {sql_literal(hn)}, {period}, {sql_literal(sbp_value(values.get('vitalsign_sbp')))}, "
                    f"{sql_literal(dbp_value(values.get('vitalsign_dbp')))}, {sql_literal(pulse_value(values.get('vitalsign_hr')))}, "
                    f"{sql_literal(respiratory_value(values.get('vitalsign_resp')))}, {sql_literal(hba1c_value(values.get('lab_hba1c')))}, "
                    f"{sql_literal(as_float(values.get('lab_c_peptide')))}, {sql_literal(as_float(values.get('lab_ldl')))}, "
                    f"{flag(row.get('med_metformin'))}, {flag(row.get('med_insulin'))}, {flag(row.get('med_arb'))}, "
                    f"{flag(row.get('med_ccb'))}, {flag(row.get('med_acei'))} "
                    f"FROM patients p WHERE p.hn={sql_literal(hn)} AND NOT EXISTS ("
                    f"SELECT 1 FROM patient_longitudinal_records r WHERE r.hn={sql_literal(hn)} AND r.period_offset={period});"
                )
    statements.append("COMMIT;")
    return "\n".join(statements) + "\n"


def write_seed_chunks(seed_sql: str, chunk_dir: Path, chunk_size: int = 350) -> None:
    chunk_dir.mkdir(parents=True, exist_ok=True)
    for old_chunk in chunk_dir.glob("chunk_*.sql"):
        old_chunk.unlink()
    inserts = [line for line in seed_sql.splitlines() if line.startswith("INSERT INTO ")]
    total = max(1, math.ceil(len(inserts) / chunk_size))
    for index in range(total):
        start = index * chunk_size
        lines = [
            f"-- USE MED agent dataset seed chunk {index + 1} of {total}",
            "BEGIN;",
            *inserts[start:start + chunk_size],
            "COMMIT;",
            "",
        ]
        (chunk_dir / f"chunk_{index + 1:02d}.sql").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate USE MED SQL seed from agent Excel datasets.")
    parser.add_argument("--dataset-dir", type=Path, default=DEFAULT_DATASET_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--chunk-dir", type=Path, default=DEFAULT_CHUNK_DIR)
    parser.add_argument("--chunk-size", type=int, default=350)
    parser.add_argument("--no-chunks", action="store_true")
    args = parser.parse_args()

    sql = build_sql(args.dataset_dir)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(sql, encoding="utf-8")
    if not args.no_chunks:
        write_seed_chunks(sql, args.chunk_dir, args.chunk_size)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
